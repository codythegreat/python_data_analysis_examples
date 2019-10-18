import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pprint
import re
import numpy as np

# load our excel file into python. create a list of worksheets from the book.
wb = load_workbook(filename = 'Football-2019.xlsx')
workSheets = wb.sheetnames[:5]

# fill teams list with lists of [team1, team2, winner]
teams = []
for ws in workSheets:
    for row in range(2, 16):
        gameData = wb[ws]['A'+str(row)].value.split(' at ')
        gameData.append(wb[ws]['M'+str(row)].value)
        teams.append(gameData)

# graph for teams by appearance
allTeams = []
for t in teams:
    allTeams.extend(t[:1])
x, y = np.unique(allTeams, return_counts=True)
for i in range(1, int(len(allTeams)/10)):
    plt.figure()
    # plot 10 at a time for readability
    plt.plot(x[(i-1)*10:i*10], y[(i-1)*10:i*10])
    plt.show(block=False)
# or simply print to console the team counts
for i in range(len(x)):
    print(x[i].ljust(20), "-".ljust(5), y[i], " times")

# graph for teams that won the most
# This data would require on to fill in row M with winnning team names.
winningTeams = []
for t in teams:
    winningTeams.append(t[len(t)-1])
x, y = np.unique(winningTeams, return_counts=True)
for i in range(len(x)):
    print(x[i].ljust(20), "-".ljust(5), y[i], " wins")

# fill players list from spreadsheet
players = []
for col in range(2, 13):
    # Grab the first three letters of the name in upper case
    players.append(wb.active.cell(column=col, row=1).value[:3].upper())

# create a new list holding weeks > players > scores
winsByWeekByPlayer = []
# add 5 empty lists to hold each weeks game data 
for week in range(0, 5):
    winsByWeekByPlayer.append([])

for ws in workSheets:
    colCounter = 2
    for player in players:
        # new list to hold each pick score
        scoreByGame = []
        for row in range(2, 19):
            # since some weeks are shorter, break when reach score
            if re.search(r"\d+-\d+", wb[ws].cell(row=row, column=colCounter).value):
                break
            # determine win by fill color
            if wb[ws].cell(row=row, column=colCounter).fill.fill_type == 'solid':
                scoreByGame.append(0)
            elif wb[ws].cell(row=row, column=colCounter).fill.fill_type == None:
                scoreByGame.append(1)
        colCounter = colCounter + 1
        winsByWeekByPlayer[workSheets.index(ws)].append(scoreByGame)

# use np.sum to sum each player's list of scores to get a total score we can plot
plt.figure()
for p in range(0, len(players)):
    playerValues = []
    for week in winsByWeekByPlayer:
        playerValues.append(np.sum(week[p]))
    label = players[p]
    plt.plot([1,2,3,4,5], playerValues, label=label)

# create a legend on the plot and show the plot
plt.legend()
plt.show()