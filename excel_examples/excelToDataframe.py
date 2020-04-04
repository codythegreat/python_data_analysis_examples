# excelToDataframe.py - reads from an existing excel 
# workbook and showcases various methods for reading,
# printing, and filtering it's data.

from openpyxl import load_workbook
import pandas as pd

# we can load an existing workbook with the load_workbook function
# the filename parameter represents the path to the file
wb = load_workbook(filename = 'SampleData.xlsx')

# we can print out the sheet titles in our loaded wb like this
for sheet in wb:
    print(sheet.title)

# we can assign a sheet to a variable like this
ws = wb["SalesOrders"]

# we can then access individual cells within the worksheet
cell = ws['A1']

# to access the value at this cell, we can use the value method
print(cell.value)

# you can also use integer values to find a cell
cell2 = ws.cell(row=1, column=2)
print(cell2.value)

# When you create a cell like this it is stored in memory
# so to be efficient if you want to display a large amount
# of data it is better to user iter_rows() and iter_cols()
for row in ws.iter_rows(min_row=1, max_col=5, max_row=10):
    for cell in row:
        print('|' + str.center(str(cell.value), 25), end="")
    print('|')


# We can also convert out excel sheet to a DataFrame
data = ws.values

# get the first line in the sheet as the header
columns = next(data)[0:]

# create your DataFrame
df = pd.DataFrame(data, columns=columns)

# you can print the DataFrame to get the entire sheet's values
print(df)
print()

# of course with a DataFrame we have access to all of panda's
# DataFrame class methods. This allows us to do things like filter
# here we've filtered down to rows where Jones is selling
# either Pens or Pen Sets
print(df.loc[df['Item'].isin(['Pen', 'Pen Set'])]
    .loc[df['Rep'] == 'Jones'])